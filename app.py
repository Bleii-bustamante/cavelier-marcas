import os
import re
import unicodedata
import json
import secrets
from io import BytesIO
from difflib import SequenceMatcher

# Librerías optimizadas para memoria
import openpyxl
from PIL import Image
from fpdf import FPDF
from fpdf.enums import XPos, YPos
from flask import (Flask, render_template, request, redirect,
                   url_for, session, send_file, jsonify, flash)

# ─────────────────────────────────────────────────────────────
# CONFIGURACIÓN
# ─────────────────────────────────────────────────────────────
app = Flask(__name__)
app.secret_key = secrets.token_hex(32)

UMBRAL_CORTE   = 60
UPLOAD_FOLDER  = "uploads"
PDF_FOLDER     = "pdfs_generados"
CARPETA_IMG    = "temp_logos"
TEXTO_PIE      = "Cra. 4 N° 72A - 35 Bogotá D.C. | Tel. (+57) 601 3473611 | cavelier@cavelier.com"

# Asegurar que existan las carpetas necesarias
for folder in [UPLOAD_FOLDER, PDF_FOLDER, CARPETA_IMG]:
    os.makedirs(folder, exist_ok=True)

# Usuarios autorizados
USUARIOS = {
    "cavelier": "marcas2024",
    "abogado1": "clave123",
}

# Tabla de Clases Vinculadas
CLASES_VINCULADAS = [
    {1, 2}, {1, 5}, {1, 6}, {1, 16}, {1, 17}, {1, 19}, {1, 31}, {1, 40},
    {2, 4}, {2, 16}, {2, 19},
    {3, 5}, {3, 10}, {3, 14}, {3, 18}, {3, 21}, {3, 25}, {3, 44},
    {4, 12}, {4, 13}, {4, 37}, {4, 39}, {4, 40},
    {5, 10}, {5, 29}, {5, 30}, {5, 31}, {5, 44},
    {6, 8}, {6, 19}, {7, 8}, {7, 12}, {7, 17}, {8, 21},
    {9, 10}, {9, 15}, {9, 16}, {9, 28}, {9, 35}, {9, 38}, {9, 41}, {9, 42}, {9, 45},
    {10, 44}, {11, 19}, {11, 21}, {11, 37}, {12, 37}, {12, 39}, {12, 40},
    {14, 18}, {14, 25}, {15, 41}, {16, 17}, {16, 35}, {16, 38}, {16, 41},
    {18, 25}, {19, 27}, {19, 37}, {20, 21}, {20, 22}, {20, 24}, {20, 27},
    {21, 24}, {22, 24}, {22, 27}, {22, 28}, {23, 24}, {23, 26}, {24, 25}, {24, 26},
    {26, 31}, {28, 41}, {29, 30}, {29, 31}, {29, 32}, {29, 43},
    {30, 31}, {30, 32}, {30, 43}, {31, 32}, {31, 43}, {31, 44},
    {32, 33}, {32, 43}, {33, 34}, {33, 43}, {35, 36}, {35, 38}, {35, 41}, {35, 42},
    {36, 37}, {37, 40}, {38, 41}, {38, 42}, {39, 43},
]

# ─────────────────────────────────────────────────────────────
# UTILIDADES LIGERAS
# ─────────────────────────────────────────────────────────────
def safe_text(text):
    if text is None or str(text).strip().lower() in ("", "nan", "none"):
        return "N/A"
    t = str(text).replace('\u2018',"'").replace('\u2019',"'").replace('\u201c','"')\
                .replace('\u201d','"').replace('\u2013','-')
    return t.encode('latin-1','replace').decode('latin-1')

def limpiar_titular(texto):
    if texto is None: return "Sin Titular"
    t = str(texto).upper()
    partes = re.split(r';| Y ', t)
    patrones = [r", CL",r", KR",r", CR",r", CRA",r", CALLE",r", CARRERA",
                r", AVENIDA",r", AV\.",r", EDIFICIO",r", APTO",r", BOGOTA"]
    limpios = []
    for p in partes:
        tmp = p.strip()
        for pat in patrones:
            tmp = re.split(pat, tmp)[0]
        limpios.append(tmp.strip())
    res = " / ".join(filter(None, limpios))
    return res if res else t.strip()

def limpiar(texto):
    if texto is None or str(texto).strip() == "": return ""
    texto = str(texto).upper().strip()
    if len(texto) <= 1: return ""
    texto = unicodedata.normalize('NFD', texto).encode('ascii','ignore').decode('utf-8')
    return re.sub(r'[^A-Z0-9 ]','', texto)

def limpiar_id(texto):
    return re.sub(r'[^A-Z0-9]','_', str(texto).upper())

def calcular_similitud(a, b):
    if not a or not b: return 0
    ratio = SequenceMatcher(None, a, b).ratio()
    if a in b or b in a: ratio = max(ratio, 0.85)
    return round(ratio * 100, 2)

def extraer_clases(texto):
    if texto is None or str(texto).strip() == "": return set()
    return set(int(x) for x in re.findall(r'\d+', str(texto)))

def clases_en_conflicto(cc, cg):
    for a in cc:
        for b in cg:
            if a == b: return True
            for par in CLASES_VINCULADAS:
                if a in par and b in par: return True
    return False

def calcular_clases_conflicto(cc, cg):
    resultado = set()
    for a in cc:
        for b in cg:
            if a == b:
                resultado.add(a); resultado.add(b)
            for par in CLASES_VINCULADAS:
                if a in par and b in par:
                    resultado.add(a); resultado.add(b)
    return resultado

def filtrar_productos(texto, clases):
    if texto is None or str(texto).strip() == "": return "N/A"
    bloques = re.split(r'(?=\b\d{1,2}\.\s)', str(texto).strip())
    resultado = []
    for bloque in bloques:
        bloque = bloque.strip()
        if not bloque: continue
        m = re.match(r'^(\d{1,2})\.\s', bloque)
        if m and int(m.group(1)) in clases:
            resultado.append(bloque)
        elif not m:
            resultado.append(bloque)
    return " ".join(resultado) if resultado else str(texto)

# ─────────────────────────────────────────────────────────────
# LECTURA DE EXCEL (OPENPYXL)
# ─────────────────────────────────────────────────────────────
def leer_excel_sin_pandas(file_bytes, carpeta_logos):
    bio = BytesIO(file_bytes)
    wb = openpyxl.load_workbook(bio, data_only=True)
    sheet = wb.active

    # Extraer Logos
    if hasattr(sheet, '_images'):
        for img in sheet._images:
            try:
                row = img.anchor._from.row + 1
                exp_id = sheet.cell(row=row, column=1).value
                if exp_id:
                    ruta = os.path.join(carpeta_logos, f"{limpiar_id(exp_id)}.png")
                    with open(ruta, "wb") as f: f.write(img._data())
            except: continue

    datos = []
    # Itera filas evitando cargar todo a la vez
    for row_cells in sheet.iter_rows(min_row=2, values_only=True):
        if len(row_cells) < 2 or not row_cells[1]: continue
        
        marca_orig = str(row_cells[1])
        marca_limp = limpiar(marca_orig)
        if not marca_limp: continue

        # Formatear fecha
        fecha_val = row_cells[3]
        fecha_str = fecha_val.strftime('%d/%m/%Y') if hasattr(fecha_val, 'strftime') else str(fecha_val)

        datos.append({
            "Expediente_ID": str(row_cells[0]),
            "Fecha_Rad": fecha_str,
            "Titular": limpiar_titular(row_cells[7]),
            "Marca_Original": marca_orig,
            "Marca_Limpia": marca_limp,
            "Productos_Texto": str(row_cells[10]) if len(row_cells) > 10 else "N/A",
            "Clases": extraer_clases(row_cells[8])
        })
    return datos

# ─────────────────────────────────────────────────────────────
# GENERACIÓN DE PDF
# ─────────────────────────────────────────────────────────────
class PDFCavelier(FPDF):
    def header(self):
        if os.path.exists("encabezado.png"):
            self.image("encabezado.png", x=60, y=8, w=90)
        self.ln(25)
    def footer(self):
        self.set_y(-15)
        self.set_text_color(44,62,80)
        self.set_font("Helvetica","B",7)
        self.set_draw_color(44,62,80)
        self.line(10,282,200,282)
        self.cell(0,10,f"{TEXTO_PIE}  |  Página {self.page_no()}", align="C")

def generar_pdf(c, g, score, clases_conf, concepto="", ruta_pdf=None):
    pdf = PDFCavelier()
    pdf.add_page()
    pdf.set_font("Helvetica","B",16); pdf.set_text_color(44,62,80)
    pdf.cell(190,10,"INFORME TÉCNICO DE OPOSICIÓN", new_x=XPos.LMARGIN, new_y=YPos.NEXT, align="C")
    pdf.ln(2)

    # Barra Similitud
    pdf.set_fill_color(240,240,240); pdf.rect(55, pdf.get_y(), 80, 4, style="F")
    color = (231,76,60) if score>=80 else (241,196,15) if score>=70 else (46,204,113)
    pdf.set_fill_color(*color); pdf.rect(55, pdf.get_y(), score*0.8, 4, style="F")
    pdf.set_font("Helvetica","B",10); pdf.set_x(140)
    pdf.cell(30,4,f"{score}% Similitud", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.ln(10)

    # Logos
    y_img = pdf.get_y()
    def colocar_logo(exp_id, x_marco):
        ruta = os.path.join(CARPETA_IMG, f"{limpiar_id(exp_id)}.png")
        if os.path.exists(ruta):
            pdf.set_draw_color(220,220,220); pdf.rect(x_marco, y_img, 50, 40)
            try:
                with Image.open(ruta) as im:
                    im = im.convert("RGBA")
                    # Redimensión básica para evitar errores de memoria en PDF
                    im.thumbnail((400, 400))
                    tmp_p = ruta.replace(".png","_tmp.png")
                    im.save(tmp_p)
                    pdf.image(tmp_p, x=x_marco+5, y=y_img+5, w=40)
                    os.remove(tmp_p)
            except: pass

    colocar_logo(c['Expediente_ID'], 25)
    colocar_logo(g['Expediente_ID'], 115)
    pdf.set_y(y_img + 45)

    # Datos Comparativos
    pdf.set_font("Helvetica","B",11); pdf.set_text_color(0,0,0)
    pdf.cell(95,7,f"CLIENTE: {safe_text(c['Marca_Original'])}", align="L")
    pdf.cell(95,7,f"GACETA: {safe_text(g['Marca_Original'])}", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    
    pdf.set_font("Helvetica","",9); pdf.set_text_color(80,80,80)
    pdf.cell(95,5,f"Exp: {c['Expediente_ID']}", align="L")
    pdf.cell(95,5,f"Exp: {g['Expediente_ID']}", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    
    # Productos filtrados
    pdf.ln(5); pdf.set_font("Helvetica","B",10); pdf.set_text_color(44,62,80)
    pdf.cell(95,7," PRODUCTOS (CLIENTE):", border="B", align="L")
    pdf.cell(95,7," PRODUCTOS (GACETA):", border="B", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    
    pdf.set_font("Helvetica","",8); pdf.set_text_color(0,0,0)
    y_antes_prod = pdf.get_y()
    pdf.multi_cell(90, 4, safe_text(filtrar_productos(c['Productos_Texto'], clases_conf)), align="J")
    y_fin_c = pdf.get_y()
    pdf.set_y(y_antes_prod); pdf.set_x(105)
    pdf.multi_cell(90, 4, safe_text(filtrar_productos(g['Productos_Texto'], clases_conf)), align="J")
    y_fin_g = pdf.get_y()
    
    # Concepto Jurídico
    pdf.set_y(max(y_fin_c, y_fin_g) + 10)
    if concepto:
        pdf.set_font("Helvetica","B",10); pdf.set_text_color(44,62,80)
        pdf.cell(190,7," CONCEPTO JURÍDICO:", border="B", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        pdf.set_font("Helvetica","",9); pdf.set_text_color(0,0,0)
        pdf.multi_cell(190, 5, safe_text(concepto), align="J")

    if ruta_pdf:
        pdf.output(ruta_pdf)
    return pdf

# ─────────────────────────────────────────────────────────────
# RUTAS FLASK
# ─────────────────────────────────────────────────────────────
def login_requerido(f):
    from functools import wraps
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not session.get("usuario"): return redirect(url_for("login"))
        return f(*args, **kwargs)
    return wrapper

@app.route("/", methods=["GET","POST"])
def login():
    if request.method == "POST":
        u, p = request.form.get("usuario"), request.form.get("password")
        if USUARIOS.get(u) == p:
            session["usuario"] = u
            return redirect(url_for("inicio"))
        flash("Credenciales incorrectas")
    return render_template("login.html")

@app.route("/inicio")
@login_requerido
def inicio():
    return render_template("inicio.html", usuario=session["usuario"])

import gc # Importante para liberar RAM manualmente

# ... (Manten tus imports y utilidades de limpieza)

import gc

@app.route("/procesar", methods=["POST"])
@login_requerido
def procesar():
    file_c = request.files.get("archivo_clientes")
    file_g = request.files.get("archivo_gaceta")
    
    if not file_c or not file_g:
        flash("Faltan archivos")
        return redirect(url_for("inicio"))

    # 1. Limpiar carpetas
    for folder in [CARPETA_IMG, PDF_FOLDER]:
        for f in os.listdir(folder):
            try: os.remove(os.path.join(folder, f))
            except: pass

    # Guardar archivos temporalmente en disco para leerlos fila por fila
    path_c = os.path.join(UPLOAD_FOLDER, "clientes.xlsx")
    path_g = os.path.join(UPLOAD_FOLDER, "gaceta.xlsx")
    file_c.save(path_c)
    file_g.save(path_g)

    def iterar_excel(path):
        """Generador que lee el excel fila por fila sin cargar todo a RAM"""
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheet = wb.active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 2 or not row[1]: continue
            marca_orig = str(row[1])
            marca_limp = limpiar(marca_orig)
            if not marca_limp: continue
            
            yield {
                "Expediente_ID": str(row[0]),
                "Marca_Original": marca_orig,
                "Marca_Limpia": marca_limp,
                "Clases": extraer_clases(row[8]) if len(row) > 8 else set(),
                "Titular": str(row[7]) if len(row) > 7 else "",
                "Productos_Texto": str(row[10]) if len(row) > 10 else ""
            }
        wb.close()

    resultados = []
    
    # 2. PROCESAMIENTO SECUENCIAL (MARCA POR MARCA)
    # Leemos clientes uno por uno
    for c in iterar_excel(path_c):
        clases_c = c["Clases"]
        m_c = c["Marca_Limpia"]

        # Por cada cliente, abrimos y recorremos la gaceta
        # Esto es más lento pero usa CASI CERO RAM
        for g in iterar_excel(path_g):
            if clases_en_conflicto(clases_c, g["Clases"]):
                score = calcular_similitud(m_c, g["Marca_Limpia"])
                
                if score >= UMBRAL_CORTE:
                    p_id = f"{limpiar_id(c['Marca_Original'])}_vs_{limpiar_id(g['Marca_Original'])}"
                    
                    resultados.append({
                        "pdf_id": p_id,
                        "marca_cliente": c["Marca_Original"],
                        "score": score,
                        "marca_gaceta": g["Marca_Original"],
                        "_full_c": {k: (list(v) if isinstance(v, set) else v) for k, v in c.items()},
                        "_full_g": {k: (list(v) if isinstance(v, set) else v) for k, v in g.items()},
                        "_clases_conf": list(calcular_clases_conflicto(clases_c, g["Clases"]))
                    })
        
        # Limpieza manual después de procesar cada cliente
        gc.collect()

    # 3. EXTRAER LOGOS SOLO DE LOS RESULTADOS (Para ahorrar RAM)
    if resultados:
        # Abrimos los excels una última vez solo para sacar las fotos de los que sí cruzaron
        def extraer_logos_finales(path, ids_necesarios):
            wb = openpyxl.load_workbook(path, data_only=True)
            sheet = wb.active
            if hasattr(sheet, '_images'):
                for img in sheet._images:
                    row_idx = img.anchor._from.row + 1
                    exp_id = str(sheet.cell(row=row_idx, column=1).value)
                    if exp_id in ids_necesarios:
                        ruta = os.path.join(CARPETA_IMG, f"{limpiar_id(exp_id)}.png")
                        with open(ruta, "wb") as f: f.write(img._data())
            wb.close()

        ids_c = {r["_full_c"]["Expediente_ID"] for r in resultados}
        ids_g = {r["_full_g"]["Expediente_ID"] for r in resultados}
        extraer_logos_finales(path_c, ids_c)
        extraer_logos_finales(path_g, ids_g)

    # 4. Finalizar
    if not resultados:
        flash("No se encontraron conflictos.")
        return redirect(url_for("inicio"))

    session["resultados_meta"] = [
        {"pdf_id": r["pdf_id"], "marca_cliente": r["marca_cliente"], 
         "score": r["score"], "marca_gaceta": r["marca_gaceta"]} 
        for r in resultados
    ]
    
    with open("datos_sesion.json", "w", encoding="utf-8") as f:
        json.dump(resultados, f, ensure_ascii=False)

    return redirect(url_for("resultados"))
    
    # 3. Comparación Optimizada
    for c in data_c:
        m_c = c["Marca_Limpia"]
        clases_c = c["Clases"]
        
        # Filtro de seguridad: si la marca está vacía o es muy corta, saltar
        if len(m_c) < 2: continue

        for g in data_g:
            m_g = g["Marca_Limpia"]
            
            # FILTRO RÁPIDO (Cero costo de RAM): 
            # Si no comparten clases vinculadas, ni siquiera calculamos similitud
            if not clases_en_conflicto(clases_c, g["Clases"]):
                continue

            # FILTRO DE TEXTO: Si no comparten al menos la primera letra o una es subcadena
            # esto evita llamar a SequenceMatcher innecesariamente
            if m_c[0] != m_g[0] and (m_c not in m_g and m_g not in m_c):
                # Solo calculamos similitud si son marcas "prometedoras"
                # o puedes comentar este 'if' si necesitas precisión total, pero gasta más RAM
                pass 

            score = calcular_similitud(m_c, m_g)
            
            if score >= UMBRAL_CORTE:
                p_id = f"{limpiar_id(c['Marca_Original'])}_vs_{limpiar_id(g['Marca_Original'])}"
                
                # Guardamos solo lo estrictamente necesario
                resultados.append({
                    "pdf_id": p_id,
                    "marca_cliente": c["Marca_Original"],
                    "score": score,
                    "marca_gaceta": g["Marca_Original"],
                    "concepto": "",
                    # Guardamos la data completa solo para reconstruir el PDF luego
                    "_full_c": {k: (list(v) if isinstance(v, set) else v) for k, v in c.items()},
                    "_full_g": {k: (list(v) if isinstance(v, set) else v) for k, v in g.items()},
                    "_clases_conf": list(calcular_clases_conflicto(clases_c, g["Clases"]))
                })
        
        # Liberar memoria cada 50 clientes procesados
        if len(resultados) % 50 == 0:
            gc.collect()

    if not resultados:
        flash("No se encontraron conflictos.")
        return redirect(url_for("inicio"))

    # 4. Guardar y limpiar sesión
    session["resultados_meta"] = [
        {"pdf_id": r["pdf_id"], "marca_cliente": r["marca_cliente"], 
         "score": r["score"], "marca_gaceta": r["marca_gaceta"]} 
        for r in resultados
    ]
    
    with open("datos_sesion.json", "w", encoding="utf-8") as f:
        json.dump(resultados, f, ensure_ascii=False)

    # Limpieza final de variables grandes
    del data_c
    del data_g
    gc.collect()

    return redirect(url_for("resultados"))

    # Guardamos los metadatos ligeros en la sesión
    session["resultados_meta"] = [{k:v for k,v in r.items() if not k.startswith("_")} for r in resultados]
    
    # Guardamos los datos pesados (logos y diccionarios completos) en un archivo temporal
    with open("datos_sesion.json", "w", encoding="utf-8") as f:
        json.dump(resultados, f, ensure_ascii=False)

    return redirect(url_for("resultados"))

    # Limpiar procesos anteriores
    for folder in [PDF_FOLDER, CARPETA_IMG]:
        for f in os.listdir(folder):
            try: os.remove(os.path.join(folder, f))
            except: pass

    data_c = leer_excel_sin_pandas(file_c.read(), CARPETA_IMG)
    data_g = leer_excel_sin_pandas(file_g.read(), CARPETA_IMG)

    resultados = []
    for c in data_c:
        for g in data_g:
            score = calcular_similitud(c["Marca_Limpia"], g["Marca_Limpia"])
            if score >= UMBRAL_CORTE and clases_en_conflicto(c["Clases"], g["Clases"]):
                clases_conf = calcular_clases_conflicto(c["Clases"], g["Clases"])
                p_id = f"{limpiar_id(c['Marca_Original'])}_vs_{limpiar_id(g['Marca_Original'])}"
                
                # Crear PDF inicial
                generar_pdf(c, g, score, clases_conf, ruta_pdf=os.path.join(PDF_FOLDER, f"{p_id}.pdf"))
                
                # Preparar para JSON (convertir sets a listas)
                c_json = c.copy(); c_json["Clases"] = list(c["Clases"])
                g_json = g.copy(); g_json["Clases"] = list(g["Clases"])

                resultados.append({
                    "pdf_id": p_id,
                    "exp_cliente": c["Expediente_ID"],
                    "marca_cliente": c["Marca_Original"],
                    "titular_cliente": c["Titular"],
                    "score": score,
                    "exp_gaceta": g["Expediente_ID"],
                    "marca_gaceta": g["Marca_Original"],
                    "clases_c": str(list(c["Clases"])),
                    "clases_g": str(list(g["Clases"])),
                    "concepto": "",
                    "_full_c": c_json,
                    "_full_g": g_json,
                    "_clases_conf": list(clases_conf)
                })

    if not resultados:
        flash("No se encontraron conflictos.")
        return redirect(url_for("inicio"))

    # Guardar en JSON para no sobrecargar la memoria de la Sesión de Flask
    session["resultados_meta"] = [{k:v for k,v in r.items() if not k.startswith("_")} for r in resultados]
    with open("datos_sesion.json", "w", encoding="utf-8") as f:
        json.dump(resultados, f, ensure_ascii=False)

    return redirect(url_for("resultados"))

@app.route("/resultados")
@login_requerido
def resultados():
    return render_template("resultados.html", resultados=session.get("resultados_meta", []))

@app.route("/guardar_concepto", methods=["POST"])
@login_requerido
def guardar_concepto():
    data = request.json
    p_id, concepto = data.get("pdf_id"), data.get("concepto")
    
    with open("datos_sesion.json", "r", encoding="utf-8") as f:
        todos = json.load(f)
    
    for item in todos:
        if item["pdf_id"] == p_id:
            item["concepto"] = concepto # Guardamos el concepto en el JSON
            break
            
    with open("datos_sesion.json", "w", encoding="utf-8") as f:
        json.dump(todos, f, ensure_ascii=False)
        
    return jsonify({"ok": True})
    item = next((x for x in todos if x["pdf_id"] == p_id), None)
    if item:
        generar_pdf(item["_full_c"], item["_full_g"], item["score"], 
                    set(item["_clases_conf"]), concepto=concepto, 
                    ruta_pdf=os.path.join(PDF_FOLDER, f"{p_id}.pdf"))
        return jsonify({"ok": True})
    return jsonify({"ok": False}), 404

@app.route("/descargar_pdf/<pdf_id>")
@login_requerido
def descargar_pdf(pdf_id):
    # Intentamos cargar los datos guardados en el archivo temporal
    try:
        with open("datos_sesion.json", "r", encoding="utf-8") as f:
            todos = json.load(f)
    except:
        return "Sesión expirada o archivos no encontrados. Por favor procesa de nuevo.", 404
    
    # Buscamos la coincidencia específica
    item = next((x for x in todos if x["pdf_id"] == pdf_id), None)
    if not item:
        return "No se encontraron datos para este PDF.", 404

    ruta_salida = os.path.join(PDF_FOLDER, f"{pdf_id}.pdf")
    
    # GENERACIÓN BAJO DEMANDA: Solo procesamos este PDF individual
    generar_pdf(
        item["_full_c"], 
        item["_full_g"], 
        item["score"], 
        set(item["_clases_conf"]), 
        concepto=item.get("concepto", ""),
        ruta_pdf=ruta_salida
    )
    
    return send_file(ruta_salida, as_attachment=True)

@app.route("/descargar_excel")
@login_requerido
def descargar_excel():
    metas = session.get("resultados_meta", [])
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Marca Cliente", "Similitud", "Marca Gaceta", "Clases Cliente", "Clases Gaceta"])
    for r in metas:
        ws.append([r["marca_cliente"], r["score"], r["marca_gaceta"], r["clases_c"], r["clases_g"]])
    
    buf = BytesIO()
    wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True, download_name="Reporte.xlsx")

if __name__ == "__main__":
    # Render asigna un puerto dinámico, lo capturamos con os.environ
    port = int(os.environ.get("PORT", 5000))
    # Importante: host="0.0.0.0" para que sea visible externamente
    app.run(host="0.0.0.0", port=port, debug=False)
